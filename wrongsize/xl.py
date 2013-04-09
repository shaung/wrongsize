# coding: utf-8

import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import get_column_letter

import sys
import sqlalchemy

from attrlist import attrs


def fmt(s):
    if s == 'NULL':
        return None
    return s


def encode(v, coding='utf8'):
    if isinstance(v, unicode):
        try:
            return v.encode(coding)
        except:
            return v.encode('cp932')
    elif isinstance(v, str):
        return v.decode('cp932').encode('utf8')
    elif v is None:
        return 'NULL'
    else:
        return str(v)


def create_data(sht, table, cond_case=None):
    with_col_desc = (sht.cell(row=0, column=0).value is None)
    rowpos = with_col_desc and 2 or 0
    head = [sht.cell(row=rowpos, column=c).value for c in range(254)]
    head = [x for x in head if x is not None]
    sht.title = table.get_short_name() + sht.title

    for r in range(rowpos + 1, 10000):
        if sht.cell(row=r, column=3).value is None:
            break
        if sht.cell(row=r, column=0).value is None:
            continue
        row = [fmt(sht.cell(row=r, column=c).value) for c in range(len(head))]
        data = dict(zip(head, row))
        if cond_case is not None and not cond_case(data):
            continue

        columns = table.get_colname_list()
        for col in data.keys():
            if not col in columns:
                del data[col]

        try:
            table.new(data)
        except sqlalchemy.exc.IntegrityError, e:
            print >> sys.stderr, e
            print >> sys.stderr, data
            pass
        

def write_data_sheet(sht, table, cond=None, with_col_desc=False):
    result = table.select(cond).execute()
    coltypes = dict(((c.name, c.type) for c in table.columns))

    rowpos = with_col_desc and 2 or 0
    if with_col_desc:
        for c, colname in enumerate(result.keys()):
            coldesc = attrs.get(colname, '')
            cell = sht.cell(row=0, column=c)
            cell.value = encode(coldesc)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = 'EFEFEF'

            coltype = coltypes[colname]
            cell = sht.cell(row=1, column=c)
            cell.value = encode(coltype).lower()
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = 'EFEFEF'

    for c, colname in enumerate(result.keys()):
        cell = sht.cell(row=rowpos, column=c)
        cell.value = encode(colname if not table.prefix else colname[7:])
        cell.style.fill.fill_type = Fill.FILL_SOLID
        cell.style.fill.start_color.index = 'EFEFEF'

    for r, row in enumerate(result.fetchall(), start=rowpos+1):
        for c, (k, v) in enumerate(row.items()):
            cell = sht.cell(row=r, column=c)
            #cell.value = encode(v)
            cell.set_value_explicit(value=encode(v))
            #cell.style.fill.fill_type = Fill.FILL_SOLID
            #cell.style.fill.start_color.index = Color.YELLOW

def compare_all(wb, with_desc=False):
    shtnames = wb.get_sheet_names()
    rslts = []
    for name in shtnames:
        name_new = name + u' (後)'
        if name_new not in shtnames:
            #print 'not table', name.encode('cp932')
            continue
        sht_old = wb.get_sheet_by_name(name)
        sht_new = wb.get_sheet_by_name(name_new)
        short_name = name[:4].lower()
        table = globals().get(short_name)
        if not table:
            #print 'not found', short_name.encode('cp932')
            continue
        rslt = compare_sheets(table, sht_old, sht_new, with_desc)
        rslts.append((name, rslt))
    return rslts


def compare_sheets(table, sht_old, sht_new, with_desc=False):
    rslt = {
        'tablename' : table.table_name,
        'newdata'   : [],
        'moddata'   : [],
    }
    cols = table.get_colname_list()
    pks = insp.get_pk_constraint(table.table_name)['constrained_columns']
    pks = [x[len(table.prefix):] for x in pks]
    header, data_old = get_all_sheet_data(sht_old, pks, with_desc)
    _, data_new = get_all_sheet_data(sht_new, pks, with_desc)
    # new data
    new_keys = [k for k in data_new if k not in data_old]
    for key in new_keys:
        row, data = data_new[key]
        rslt['newdata'].append((key, dict((k + ' ' + attrs.get(table.prefix + k, ''), v) for (k, v) in data.iteritems())))
        for colname in data:
            col = header[colname]
            cell = sht_new.cell(row=row, column=col)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.YELLOW
    # deleted data
    # changed data
    mod_keys = [k for k in data_new if k in data_old]
    for key in mod_keys:
        rold, dold = data_old[key]
        rnew, dnew = data_new[key]
        """
        print table.table_name
        print 'added', set(dnew.keys()) - set(dold.keys())
        print 'removed', set(dold.keys()) - set(dnew.keys())
        """
        moddata = {}
        for colname in (k for k in dnew if dnew.get(k) != dold.get(k)):
            col = header[colname]

            cell = sht_old.cell(row=rold, column=col)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.YELLOW

            cell = sht_new.cell(row=rnew, column=col)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.YELLOW
            coldesc = attrs.get(table.prefix + colname, '')
            moddata[colname + ' ' + coldesc] = {
                'old' : dold.get(colname),
                'new' : dnew.get(colname),
            }
        if moddata:
            rslt['moddata'].append((k, moddata))
    return rslt


def get_all_sheet_data(sht, pks, with_desc=False):
    with_col_desc = with_desc or (sht.cell(row=0, column=0).value is None)
    rowpos = with_col_desc and 2 or 0

    result = {}

    header = [(c, sht.cell(row=rowpos, column=c).value) for c in range(254)]
    header = [x for x in header if x[-1] is not None]
    head = [x for (c, x) in header]
    header = dict(((name, col) for col, name in header))

    for r in xrange(rowpos + 1, 10000):
        if sht.cell(row=r, column=3).value is None:
            break
        if sht.cell(row=r, column=0).value is None:
            continue
        row = [fmt(sht.cell(row=r, column=c).value) for c in range(len(head))]
        data = dict(zip(head, row))
        try:
            key = '__and__'.join(map(str, ['%s=%s' % (k, data[k]) for k in pks]))
        except:
            print >> sys.stderr, pks, data
            raise
        result[key] = (r, data)
    return header, result
